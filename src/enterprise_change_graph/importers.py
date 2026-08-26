from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Iterable

from .io import load_payload
from .model import EnterpriseGraph, GraphValidationError


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _metadata(value: Any, context: str) -> dict[str, Any]:
    if value is None or value == "":
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError as exc:
        raise GraphValidationError([f"{context} metadata is not valid JSON: {exc}"]) from exc
    if not isinstance(parsed, dict):
        raise GraphValidationError([f"{context} metadata must decode to an object"])
    return parsed


def _read_csv(path: str | Path) -> list[dict[str, str]]:
    source = Path(path)
    if not source.exists():
        raise GraphValidationError([f"CSV file does not exist: {source}"])
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise GraphValidationError([f"CSV file has no header: {source}"])
        return [{str(k).strip(): _clean(v) for k, v in row.items() if k is not None} for row in reader]


def graph_from_catalog_rows(*, node_rows: Iterable[dict[str, Any]], edge_rows: Iterable[dict[str, Any]] = (), change_rows: Iterable[dict[str, Any]] = (), provenance: str | None = None) -> EnterpriseGraph:
    nodes: list[dict[str, Any]] = []
    edges: list[dict[str, Any]] = []
    changes: list[dict[str, Any]] = []
    for index, row in enumerate(node_rows, 2):
        item: dict[str, Any] = {"id": _clean(row.get("id")), "type": _clean(row.get("type"))}
        if _clean(row.get("name")): item["name"] = _clean(row.get("name"))
        if _clean(row.get("criticality")): item["criticality"] = _clean(row.get("criticality"))
        metadata = _metadata(row.get("metadata"), f"nodes row {index}")
        if metadata: item["metadata"] = metadata
        if provenance: item["provenance"] = [provenance]
        nodes.append(item)
    for index, row in enumerate(edge_rows, 2):
        item = {"source": _clean(row.get("source")), "target": _clean(row.get("target")), "relation": _clean(row.get("relation"))}
        if _clean(row.get("propagation")): item["propagation"] = _clean(row.get("propagation"))
        metadata = _metadata(row.get("metadata"), f"edges row {index}")
        if metadata: item["metadata"] = metadata
        if provenance: item["provenance"] = [provenance]
        edges.append(item)
    for index, row in enumerate(change_rows, 2):
        raw_seeds = row.get("seeds", "")
        seeds = [str(seed).strip() for seed in raw_seeds if str(seed).strip()] if isinstance(raw_seeds, (list, tuple)) else [part.strip() for part in re.split(r"[;,]", str(raw_seeds)) if part.strip()]
        item = {"id": _clean(row.get("id")), "title": _clean(row.get("title")), "seeds": seeds}
        if _clean(row.get("description")): item["description"] = _clean(row.get("description"))
        if _clean(row.get("kind")): item["kind"] = _clean(row.get("kind"))
        metadata = _metadata(row.get("metadata"), f"changes row {index}")
        if metadata: item["metadata"] = metadata
        if provenance: item["provenance"] = [provenance]
        changes.append(item)
    return EnterpriseGraph.from_dict({"version": 1, "nodes": nodes, "edges": edges, "changes": changes})


def import_catalog_csv(nodes_path: str | Path, *, edges_path: str | Path | None = None, changes_path: str | Path | None = None) -> EnterpriseGraph:
    nodes_source = Path(nodes_path); node_rows = _read_csv(nodes_source); edge_rows = _read_csv(edges_path) if edges_path else []; change_rows = _read_csv(changes_path) if changes_path else []
    sources = [nodes_source.as_posix()]
    if edges_path: sources.append(Path(edges_path).as_posix())
    if changes_path: sources.append(Path(changes_path).as_posix())
    graph = graph_from_catalog_rows(node_rows=node_rows, edge_rows=edge_rows, change_rows=change_rows, provenance=";".join(sources)); graph.metadata["imported_from"] = sources; return graph


def import_catalog_workbook(path: str | Path) -> EnterpriseGraph:
    source = Path(path)
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise GraphValidationError(["Excel import requires the optional dependency: pip install enterprise-change-graph[excel]"]) from exc
    if not source.exists(): raise GraphValidationError([f"workbook does not exist: {source}"])
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    def rows(sheet_name: str) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames: return []
        values = list(workbook[sheet_name].iter_rows(values_only=True))
        if not values: return []
        headers = [_clean(value) for value in values[0]]
        return [{headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]} for row in values[1:] if any(value not in (None, "") for value in row)]
    node_rows = rows("Nodes") or rows("nodes")
    if not node_rows: raise GraphValidationError(["workbook must contain a Nodes sheet with at least one data row"])
    graph = graph_from_catalog_rows(node_rows=node_rows, edge_rows=rows("Edges") or rows("edges"), change_rows=rows("Changes") or rows("changes"), provenance=source.as_posix()); graph.metadata["imported_from"] = [source.as_posix()]; return graph


def _slug(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip()).strip("-")
    return value.lower() or "unknown"


def import_interface_as_code(path: str | Path) -> EnterpriseGraph:
    source = Path(path); payload = load_payload(source); interface = payload.get("interface")
    if not isinstance(interface, dict): raise GraphValidationError(["Interface-as-Code document must contain an 'interface' object"])
    interface_id = _clean(interface.get("id"))
    if not interface_id: raise GraphValidationError(["interface.id must be a non-empty string"])
    name = _clean(interface.get("name")) or interface_id; criticality = _clean(interface.get("criticality")) or "medium"; src = interface.get("source", {}); tgt = interface.get("target", {})
    if not isinstance(src, dict) or not isinstance(tgt, dict): raise GraphValidationError(["interface.source and interface.target must be objects"])
    src_system = _clean(src.get("system")); tgt_system = _clean(tgt.get("system"))
    if not src_system or not tgt_system: raise GraphValidationError(["interface source.system and target.system are required"])
    src_object = _clean(src.get("object")); tgt_object = _clean(tgt.get("object")); iid = f"interface.{_slug(interface_id)}"; ssid = f"system.{_slug(src_system)}"; tsid = f"system.{_slug(tgt_system)}"
    nodes: list[dict[str, Any]] = [{"id": iid, "type": "interface", "name": name, "criticality": criticality}, {"id": ssid, "type": "system", "name": src_system, "criticality": criticality}, {"id": tsid, "type": "system", "name": tgt_system, "criticality": criticality}]
    edges: list[dict[str, Any]] = [{"source": ssid, "target": iid, "relation": "provides", "propagation": "both"}, {"source": iid, "target": tsid, "relation": "delivers-to"}]
    if src_object:
        soid = f"data.{_slug(src_system)}.{_slug(src_object)}"; nodes.append({"id": soid, "type": "data", "name": f"{src_system} / {src_object}", "criticality": criticality}); edges.extend([{"source": ssid, "target": soid, "relation": "contains", "propagation": "none"}, {"source": soid, "target": iid, "relation": "published-by"}])
    if tgt_object:
        toid = f"data.{_slug(tgt_system)}.{_slug(tgt_object)}"; nodes.append({"id": toid, "type": "data", "name": f"{tgt_system} / {tgt_object}", "criticality": criticality}); edges.extend([{"source": iid, "target": toid, "relation": "writes"}, {"source": toid, "target": tsid, "relation": "stored-in", "propagation": "none"}])
    mapping = payload.get("mapping")
    if isinstance(mapping, dict) and _clean(mapping.get("profile")):
        profile = _clean(mapping.get("profile")); mid = f"mapping.{_slug(profile)}"; nodes.append({"id": mid, "type": "mapping", "name": profile, "criticality": criticality}); edges.append({"source": mid, "target": iid, "relation": "implemented-by"})
    owners: dict[str, str] = {}; ownership = payload.get("ownership")
    if isinstance(ownership, dict):
        for role in ("business", "technical", "support"):
            value = _clean(ownership.get(role))
            if value: owners[role] = value
    monitoring = payload.get("monitoring")
    if isinstance(monitoring, dict) and _clean(monitoring.get("owner")): owners["monitoring"] = _clean(monitoring.get("owner"))
    owner_roles: dict[str, list[str]] = {}
    for role, owner in sorted(owners.items()): owner_roles.setdefault(owner, []).append(role)
    for owner, roles in sorted(owner_roles.items()):
        oid = f"owner.{_slug(owner)}"
        if not any(node["id"] == oid for node in nodes): nodes.append({"id": oid, "type": "owner", "name": owner, "criticality": "low"})
        edges.append({"source": iid, "target": oid, "relation": "owned-by", "metadata": {"roles": sorted(roles)}})
    tests = payload.get("tests", [])
    if isinstance(tests, list):
        for raw_test in tests:
            if not isinstance(raw_test, dict) or not _clean(raw_test.get("id")): continue
            test_name = _clean(raw_test.get("description")) or _clean(raw_test.get("id")); tid = f"test.{_slug(interface_id)}.{_slug(_clean(raw_test.get('id')))}"; nodes.append({"id": tid, "type": "test", "name": test_name, "criticality": criticality}); edges.append({"source": iid, "target": tid, "relation": "verified-by"})
    deduped_nodes: dict[str, dict[str, Any]] = {}
    for item in nodes:
        item["provenance"] = [source.as_posix()]; existing = deduped_nodes.get(item["id"])
        if existing and existing != item: raise GraphValidationError([f"adapter generated conflicting node {item['id']}"])
        deduped_nodes[item["id"]] = item
    for item in edges: item["provenance"] = [source.as_posix()]
    return EnterpriseGraph.from_dict({"version": 1, "nodes": list(deduped_nodes.values()), "edges": edges, "relation_rules": {"owned-by": {"default": "forward", "change_kinds": {"owner-change": "forward"}}, "provides": {"default": "both"}}, "metadata": {"adapter": "interface-as-code", "source": source.as_posix()}})


def import_mapping_as_code(path: str | Path) -> EnterpriseGraph:
    source = Path(path); payload = load_payload(source); mapping = payload.get("mapping")
    if not isinstance(mapping, dict): raise GraphValidationError(["Mapping-as-Code document must contain a 'mapping' object"])
    mapping_id = _clean(mapping.get("id"))
    if not mapping_id: raise GraphValidationError(["mapping.id must be a non-empty string"])
    name = _clean(mapping.get("name")) or mapping_id; criticality = _clean(mapping.get("criticality")) or "medium"; mid = f"mapping.{_slug(mapping_id)}"
    nodes: list[dict[str, Any]] = [{"id": mid, "type": "mapping", "name": name, "criticality": criticality, "metadata": {"source_format": "mapping-as-code"}}]; edges: list[dict[str, Any]] = []
    for side, relation in (("source", "reads"), ("target", "writes")):
        endpoint = mapping.get(side, {})
        if not isinstance(endpoint, dict): continue
        system = _clean(endpoint.get("system")); obj = _clean(endpoint.get("object"))
        if system:
            sid = f"system.{_slug(system)}"; nodes.append({"id": sid, "type": "system", "name": system, "criticality": criticality})
            if obj:
                did = f"data.{_slug(system)}.{_slug(obj)}"; nodes.append({"id": did, "type": "data", "name": f"{system} / {obj}", "criticality": criticality}); edges.append({"source": did, "target": mid, "relation": relation, "propagation": "both"} if side == "source" else {"source": mid, "target": did, "relation": relation}); edges.append({"source": did, "target": sid, "relation": "stored-in", "propagation": "none"})
            else: edges.append({"source": sid, "target": mid, "relation": relation, "propagation": "both"} if side == "source" else {"source": mid, "target": sid, "relation": relation})
    raw_rules = mapping.get("fields", mapping.get("rules", []))
    if isinstance(raw_rules, list):
        for index, rule in enumerate(raw_rules, 1):
            if not isinstance(rule, dict): continue
            rid = _clean(rule.get("id")) or f"rule-{index}"; rule_node = f"mapping-rule.{_slug(mapping_id)}.{_slug(rid)}"; rule_name = _clean(rule.get("name")) or _clean(rule.get("target")) or rid; nodes.append({"id": rule_node, "type": "mapping-rule", "name": rule_name, "criticality": criticality, "metadata": {k: v for k, v in rule.items() if k not in {"id", "name"}}}); edges.append({"source": mid, "target": rule_node, "relation": "contains-rule"})
    deduped_nodes: dict[str, dict[str, Any]] = {}
    for item in nodes:
        item["provenance"] = [source.as_posix()]; existing = deduped_nodes.get(item["id"])
        if existing and existing != item: raise GraphValidationError([f"adapter generated conflicting node {item['id']}"])
        deduped_nodes[item["id"]] = item
    for item in edges: item["provenance"] = [source.as_posix()]
    return EnterpriseGraph.from_dict({"version": 1, "nodes": list(deduped_nodes.values()), "edges": edges, "metadata": {"adapter": "mapping-as-code", "source": source.as_posix()}})
